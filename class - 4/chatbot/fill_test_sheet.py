"""
Automates the EDGE Chatbot Test Sheet by driving the Streamlit UI at
http://localhost:8501/ with Selenium.

For each test case (rows 5..34):
  * Reads the prompt from column C
  * Types it into the chat, waits for the assistant reply
  * Writes the reply into column E, response time into column F
  * Asks the Mistral API to judge the answer, writing Pass/Fail into column G
    and a short rationale into column H

The original workbook is overwritten in-place; it is also saved after every
row so progress survives interruption.

Run with the chatbot venv:
    .\chatbot\venv\Scripts\python.exe fill_test_sheet.py

Requires MISTRAL_API_KEY (and optionally MISTRAL_MODEL) in the environment or
in a .env file next to this script.
"""

from __future__ import annotations

import json
import os
import time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from selenium import webdriver
from selenium.common.exceptions import (
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

URL = "http://localhost:8501/"
XLSX_INPUT = Path("input/EDGE_Chatbot_Test_Sheet.xlsx")
XLSX_OUTPUT = Path("output/EDGE_Chatbot_Test_Sheet.xlsx")

FIRST_ROW = 5
LAST_ROW = 34
COL_PROMPT = 3     # C
COL_EXPECTED = 4   # D
COL_ACTUAL = 5     # E
COL_TIME = 6       # F
COL_VERDICT = 7    # G
COL_NOTES = 8      # H

READY_TIMEOUT = 60        # seconds to wait for the app to load
ANSWER_TIMEOUT = 300      # seconds to wait for a single answer
INPUT_READY_TIMEOUT = 60  # wait for chat input to re-enable between prompts
STABLE_FOR = 2.0          # response must be unchanged for this long
POLL = 0.4                # DOM polling interval
INTER_PROMPT_PAUSE = 1.0  # small breather between prompts

load_dotenv()
MISTRAL_API_URL = "https://api.mistral.ai/v1/chat/completions"
MISTRAL_MODEL = os.environ.get("MISTRAL_MODEL", "mistral-small-latest")
JUDGE_TIMEOUT = 60  # seconds for the judge HTTP call

JUDGE_SYSTEM_PROMPT = (
    "You are a strict QA judge for a chatbot test. Compare an expected answer "
    "with an actual chatbot answer and decide whether the actual answer "
    "satisfies the expected meaning. Pass only when the actual answer is "
    "factually consistent with the expected answer and conveys its key "
    "information; minor wording differences are fine. "
    "Respond ONLY with a compact JSON object of the form "
    '{"verdict":"Pass"|"Fail","note":"<short reason, <=120 chars>"}.'
)


def simple_judge(expected: str, actual: str) -> tuple[str, str]:
    """Return (verdict, note) by asking the Mistral API to judge the answer."""
    if not actual:
        return "Fail", "empty response"
    if not expected:
        return "", ""

    api_key = os.environ.get("MISTRAL_API_KEY")
    if not api_key:
        return "", "MISTRAL_API_KEY not set"

    payload = {
        "model": MISTRAL_MODEL,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": JUDGE_SYSTEM_PROMPT},
            {
                "role": "user",
                "content": (
                    f"Expected answer:\n{expected}\n\n"
                    f"Actual answer:\n{actual}"
                ),
            },
        ],
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(
            MISTRAL_API_URL, json=payload, headers=headers, timeout=JUDGE_TIMEOUT
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        data = json.loads(content)
    except (requests.RequestException, ValueError, KeyError) as exc:
        return "", f"judge error: {exc}"

    verdict = str(data.get("verdict", "")).strip().capitalize()
    note = str(data.get("note", "")).strip()
    if verdict not in ("Pass", "Fail"):
        return "", f"unparsed verdict: {verdict!r}"
    return verdict, note


def get_messages(driver):
    return driver.find_elements(By.CSS_SELECTOR, '[data-testid="stChatMessage"]')


def get_chat_input(driver):
    return driver.find_element(
        By.CSS_SELECTOR, '[data-testid="stChatInput"] textarea'
    )


def read_text(el) -> str:
    """Return the full text content of an element, even if off-screen."""
    try:
        return (el.get_attribute("textContent") or "").strip()
    except StaleElementReferenceException:
        return ""


def wait_input_ready(driver, timeout: int = INPUT_READY_TIMEOUT):
    """Wait for the chat input to re-enable after Streamlit finishes its rerun."""
    def ready(drv):
        try:
            inp = get_chat_input(drv)
            return inp.is_enabled() and inp.get_attribute("disabled") in (None, "false")
        except StaleElementReferenceException:
            return False
    WebDriverWait(driver, timeout, poll_frequency=POLL).until(ready)


def wait_for_answer(driver, previous_count: int, timeout: int):
    """
    Return (element, text) once a new assistant bubble has rendered, the spinner
    has vanished, and its text has been stable for STABLE_FOR seconds.
    Robust to Streamlit re-renders and off-screen bubbles.
    """
    deadline = time.time() + timeout
    last_text = ""
    last_change = time.time()
    while time.time() < deadline:
        try:
            msgs = get_messages(driver)
            if len(msgs) < previous_count + 2:
                time.sleep(POLL)
                continue
            if driver.find_elements(By.CSS_SELECTOR, '[data-testid="stSpinner"]'):
                last_change = time.time()
                time.sleep(POLL)
                continue
            last = msgs[-1]
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'end', inline:'nearest'});", last
            )
            text = read_text(last)
            if not text:
                time.sleep(POLL)
                continue
            if text != last_text:
                last_text = text
                last_change = time.time()
            elif time.time() - last_change >= STABLE_FOR:
                return last, text
        except StaleElementReferenceException:
            pass
        time.sleep(POLL)
    raise TimeoutException("assistant response did not stabilize within timeout")


def main() -> None:
    if not XLSX_INPUT.exists():
        raise SystemExit(f"Workbook not found: {XLSX_INPUT}")

    XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_INPUT)
    ws = wb.active

    options = Options()
    options.add_argument("--window-size=1200,900")
    driver = webdriver.Chrome(options=options)
    try:
        driver.get(URL)
        WebDriverWait(driver, READY_TIMEOUT).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, '[data-testid="stChatInput"] textarea')
            )
        )

        for row in range(FIRST_ROW, LAST_ROW + 1):
            prompt = ws.cell(row=row, column=COL_PROMPT).value
            expected = ws.cell(row=row, column=COL_EXPECTED).value or ""
            if not prompt:
                continue

            prompt = str(prompt)
            print(f"[{row:>2}] {prompt}")

            wait_input_ready(driver)
            before = len(get_messages(driver))

            chat_input = get_chat_input(driver)
            chat_input.click()
            chat_input.send_keys(Keys.CONTROL, "a")
            chat_input.send_keys(Keys.DELETE)
            chat_input.send_keys(prompt)
            t0 = time.time()
            chat_input.send_keys(Keys.ENTER)

            try:
                _, actual = wait_for_answer(driver, before, ANSWER_TIMEOUT)
                elapsed = round(time.time() - t0, 2)
            except TimeoutException:
                actual = ""
                elapsed = ANSWER_TIMEOUT

            try:
                wait_input_ready(driver)
            except TimeoutException:
                pass
            time.sleep(INTER_PROMPT_PAUSE)

            if "smart_toy" in actual.lower():
                actual = actual.replace("smart_toy", "")

            verdict, note = simple_judge(str(expected), actual)

            ws.cell(row=row, column=COL_ACTUAL).value = actual
            ws.cell(row=row, column=COL_TIME).value = elapsed
            if verdict:
                ws.cell(row=row, column=COL_VERDICT).value = verdict
            if note and not ws.cell(row=row, column=COL_NOTES).value:
                ws.cell(row=row, column=COL_NOTES).value = note

            wb.save(XLSX_OUTPUT)
            print(f"     -> {verdict or '-':4} {elapsed:>6}s  {note}")

    finally:
        wb.save(XLSX_OUTPUT)
        driver.quit()


if __name__ == "__main__":
    main()
