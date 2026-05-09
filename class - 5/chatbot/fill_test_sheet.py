"""
Automates the EDGE Chatbot Test Sheet by driving the Streamlit UI at
http://localhost:8501/ with Selenium.

For each test case (rows 5..34):
  * Reads the prompt from column C
  * Types it into the chat, waits for the assistant reply
  * Writes the reply into column E, response time into column F
  * Runs DeepEval's GEval (Correctness) metric, backed by Mistral as the judge
    LLM, and writes Pass/Fail into column G plus the score and rationale into
    column H

The original workbook is overwritten in-place; it is also saved after every
row so progress survives interruption.

Run with the chatbot venv:
    .\chatbot\venv\Scripts\python.exe fill_test_sheet.py

Requires MISTRAL_API_KEY (and optionally MISTRAL_MODEL) in the environment or
in a .env file next to this script.
"""

from __future__ import annotations

import json
import os
import time
from pathlib import Path

import openpyxl
import requests
from deepeval.metrics import GEval
from deepeval.models import DeepEvalBaseLLM
from deepeval.test_case import LLMTestCase, SingleTurnParams
from dotenv import load_dotenv
from pydantic import BaseModel
from selenium import webdriver
from selenium.common.exceptions import (
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

URL = "http://localhost:8501/"
XLSX_INPUT = Path("input/EDGE_Chatbot_Test_Sheet.xlsx")
XLSX_OUTPUT = Path("output/EDGE_Chatbot_Test_Sheet.xlsx")

FIRST_ROW = 5
LAST_ROW = 34
COL_PROMPT = 3     # C
COL_EXPECTED = 4   # D
COL_ACTUAL = 5     # E
COL_TIME = 6       # F
COL_VERDICT = 7    # G
COL_NOTES = 8      # H

READY_TIMEOUT = 60        # seconds to wait for the app to load
ANSWER_TIMEOUT = 300      # seconds to wait for a single answer
INPUT_READY_TIMEOUT = 60  # wait for chat input to re-enable between prompts
STABLE_FOR = 2.0          # response must be unchanged for this long
POLL = 0.4                # DOM polling interval
INTER_PROMPT_PAUSE = 1.0  # small breather between prompts

load_dotenv()
MISTRAL_API_URL = "https://api.mistral.ai/v1/chat/completions"
MISTRAL_MODEL = os.environ.get("MISTRAL_MODEL", "mistral-small-latest")
JUDGE_TIMEOUT = 60  # seconds for the judge HTTP call
JUDGE_THRESHOLD = 0.5
JUDGE_CRITERIA = (
    "Determine whether the actual output is factually consistent with the "
    "expected output and conveys its key information. Minor wording differences "
    "are fine; the actual output is correct only when it satisfies the meaning "
    "of the expected output."
)


class MistralJudge(DeepEvalBaseLLM):
    """DeepEval-compatible wrapper around the Mistral chat API in JSON mode."""

    def __init__(self, model: str = MISTRAL_MODEL, api_key: str | None = None):
        self.model = model
        self.api_key = api_key or os.environ.get("MISTRAL_API_KEY", "")

    def load_model(self):
        return self.model

    def generate(self, prompt: str, schema: type[BaseModel]) -> BaseModel:
        payload = {
            "model": self.model,
            "temperature": 0,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Reply with a single JSON object that matches this "
                        f"JSON schema:\n{json.dumps(schema.model_json_schema())}"
                    ),
                },
                {"role": "user", "content": prompt},
            ],
        }
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        resp = requests.post(
            MISTRAL_API_URL, json=payload, headers=headers, timeout=JUDGE_TIMEOUT
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        return schema(**json.loads(content))

    async def a_generate(self, prompt: str, schema: type[BaseModel]) -> BaseModel:
        return self.generate(prompt, schema)

    def get_model_name(self) -> str:
        return self.model


_JUDGE_METRIC: GEval | None = None


def _get_judge_metric() -> GEval:
    global _JUDGE_METRIC
    if _JUDGE_METRIC is None:
        _JUDGE_METRIC = GEval(
            name="Correctness",
            criteria=JUDGE_CRITERIA,
            evaluation_params=[
                SingleTurnParams.EXPECTED_OUTPUT,
                SingleTurnParams.ACTUAL_OUTPUT,
            ],
            model=MistralJudge(),
            threshold=JUDGE_THRESHOLD,
        )
    return _JUDGE_METRIC


def simple_judge(expected: str, actual: str) -> tuple[str, str]:
    """Return (verdict, note) by running a DeepEval GEval metric."""
    if not actual:
        return "Fail", "empty response"
    if not expected:
        return "", ""
    if not os.environ.get("MISTRAL_API_KEY"):
        return "", "MISTRAL_API_KEY not set"

    test_case = LLMTestCase(
        input="",
        actual_output=actual,
        expected_output=expected,
    )
    try:
        metric = _get_judge_metric()
        metric.measure(test_case)
    except Exception as exc:  # network, schema, or DeepEval errors
        return "", f"judge error: {exc}"

    score = metric.score if metric.score is not None else 0.0
    reason = (metric.reason or "").strip().replace("\n", " ")
    verdict = "Pass" if metric.is_successful() else "Fail"
    note = f"score={score:.2f}; {reason}" if reason else f"score={score:.2f}"
    return verdict, note


def get_messages(driver):
    return driver.find_elements(By.CSS_SELECTOR, '[data-testid="stChatMessage"]')


def get_chat_input(driver):
    return driver.find_element(
        By.CSS_SELECTOR, '[data-testid="stChatInput"] textarea'
    )


def read_text(el) -> str:
    """Return the full text content of an element, even if off-screen."""
    try:
        return (el.get_attribute("textContent") or "").strip()
    except StaleElementReferenceException:
        return ""


def wait_input_ready(driver, timeout: int = INPUT_READY_TIMEOUT):
    """Wait for the chat input to re-enable after Streamlit finishes its rerun."""
    def ready(drv):
        try:
            inp = get_chat_input(drv)
            return inp.is_enabled() and inp.get_attribute("disabled") in (None, "false")
        except StaleElementReferenceException:
            return False
    WebDriverWait(driver, timeout, poll_frequency=POLL).until(ready)


def wait_for_answer(driver, previous_count: int, timeout: int):
    """
    Return (element, text) once a new assistant bubble has rendered, the spinner
    has vanished, and its text has been stable for STABLE_FOR seconds.
    Robust to Streamlit re-renders and off-screen bubbles.
    """
    deadline = time.time() + timeout
    last_text = ""
    last_change = time.time()
    while time.time() < deadline:
        try:
            msgs = get_messages(driver)
            if len(msgs) < previous_count + 2:
                time.sleep(POLL)
                continue
            if driver.find_elements(By.CSS_SELECTOR, '[data-testid="stSpinner"]'):
                last_change = time.time()
                time.sleep(POLL)
                continue
            last = msgs[-1]
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'end', inline:'nearest'});", last
            )
            text = read_text(last)
            if not text:
                time.sleep(POLL)
                continue
            if text != last_text:
                last_text = text
                last_change = time.time()
            elif time.time() - last_change >= STABLE_FOR:
                return last, text
        except StaleElementReferenceException:
            pass
        time.sleep(POLL)
    raise TimeoutException("assistant response did not stabilize within timeout")


def main() -> None:
    if not XLSX_INPUT.exists():
        raise SystemExit(f"Workbook not found: {XLSX_INPUT}")

    XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_INPUT)
    ws = wb.active

    options = Options()
    options.add_argument("--window-size=1200,900")
    driver = webdriver.Chrome(options=options)
    try:
        driver.get(URL)
        WebDriverWait(driver, READY_TIMEOUT).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, '[data-testid="stChatInput"] textarea')
            )
        )

        for row in range(FIRST_ROW, LAST_ROW + 1):
            prompt = ws.cell(row=row, column=COL_PROMPT).value
            expected = ws.cell(row=row, column=COL_EXPECTED).value or ""
            if not prompt:
                continue

            prompt = str(prompt)
            print(f"[{row:>2}] {prompt}")

            wait_input_ready(driver)
            before = len(get_messages(driver))

            chat_input = get_chat_input(driver)
            chat_input.click()
            chat_input.send_keys(Keys.CONTROL, "a")
            chat_input.send_keys(Keys.DELETE)
            chat_input.send_keys(prompt)
            t0 = time.time()
            chat_input.send_keys(Keys.ENTER)

            try:
                _, actual = wait_for_answer(driver, before, ANSWER_TIMEOUT)
                elapsed = round(time.time() - t0, 2)
            except TimeoutException:
                actual = ""
                elapsed = ANSWER_TIMEOUT

            try:
                wait_input_ready(driver)
            except TimeoutException:
                pass
            time.sleep(INTER_PROMPT_PAUSE)

            if "smart_toy" in actual.lower():
                actual = actual.replace("smart_toy", "")

            verdict, note = simple_judge(str(expected), actual)

            ws.cell(row=row, column=COL_ACTUAL).value = actual
            ws.cell(row=row, column=COL_TIME).value = elapsed
            if verdict:
                ws.cell(row=row, column=COL_VERDICT).value = verdict
            if note and not ws.cell(row=row, column=COL_NOTES).value:
                ws.cell(row=row, column=COL_NOTES).value = note

            wb.save(XLSX_OUTPUT)
            print(f"     -> {verdict or '-':4} {elapsed:>6}s  {note}")

    finally:
        wb.save(XLSX_OUTPUT)
        driver.quit()


if __name__ == "__main__":
    main()
